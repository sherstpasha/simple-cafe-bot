import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from db import get_connection
from datetime import datetime
import json


def auto_adjust_columns(file_path):
    wb = load_workbook(file_path)
    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            letter = get_column_letter(column_cells[0].column)
            sheet.column_dimensions[letter].width = max_length + 2
    wb.save(file_path)


def generate_reports(start_date=None, end_date=None):
    conn = get_connection()

    # 0) Приведение строк к datetime
    if isinstance(start_date, str):
        start_date = datetime.fromisoformat(start_date)
    if isinstance(end_date, str):
        end_date = datetime.fromisoformat(end_date)

    # 1) Фильтр периода
    date_filter = ""
    params = []
    if start_date and end_date:
        date_filter = "WHERE date(o.date) BETWEEN ? AND ?"
        params = [start_date.isoformat(), end_date.isoformat()]

    # 2) Читаем данные, включая признак сотрудника
    orders_df = pd.read_sql_query(
        f"""
        SELECT 
          o.date          AS date,
          o.username      AS username,
          o.is_staff      AS is_staff,
          i.payment_type  AS payment_type,
          i.item_name     AS item_name,
          i.price         AS base_price,
          i.addons_json   AS addons_json,
          i.addons_total  AS addons_total,
          i.row_total     AS row_total,
          o.raw_text      AS raw_text
        FROM orders o
        JOIN order_items i ON i.order_id = o.id
        {date_filter}
        ORDER BY o.date
        """,
        conn,
        params=params,
    )

    actions_df = pd.read_sql_query(
        "SELECT timestamp, action_type, payment_type, item_name, user_id, username, is_staff FROM actions_log",
        conn,
    )
    conn.close()

    if orders_df.empty:
        orders_df["is_staff"] = pd.Series(dtype=int)
    else:
        orders_df["is_staff"] = orders_df["is_staff"].fillna(0).astype(int)

    def _fmt_addons(raw):
        try:
            arr = json.loads(raw) if raw else []
        except Exception:
            return ""
        if not arr:
            return ""
        return ", ".join(f"{a.get('name','')} ({int(a.get('price',0))}₽)" for a in arr)

    def _prepare_orders_df(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return pd.DataFrame(
                columns=[
                    "Дата",
                    "Автор",
                    "Тип оплаты",
                    "Название",
                    "Базовая цена",
                    "Добавки",
                    "Сумма добавок",
                    "Сумма позиции",
                    "Запрос",
                    "Сотрудник",
                ]
            )

        df = df.copy()
        df["addons_text"] = df["addons_json"].apply(_fmt_addons)
        rename_map = {
            "date": "Дата",
            "username": "Автор",
            "payment_type": "Тип оплаты",
            "item_name": "Название",
            "base_price": "Базовая цена",
            "addons_text": "Добавки",
            "addons_total": "Сумма добавок",
            "row_total": "Сумма позиции",
            "raw_text": "Запрос",
            "is_staff": "Сотрудник",
        }
        df = df.rename(columns=rename_map)
        desired_cols = [
            "Дата",
            "Автор",
            "Тип оплаты",
            "Название",
            "Базовая цена",
            "Добавки",
            "Сумма добавок",
            "Сумма позиции",
            "Запрос",
            "Сотрудник",
        ]
        existing_cols = [c for c in desired_cols if c in df.columns]
        return df[existing_cols]

    def _write_report(df_prepared: pd.DataFrame, path: str):
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df_all = df_prepared.copy()
            total_all = df_all["Сумма позиции"].sum() if "Сумма позиции" in df_all else 0
            if not df_all.empty:
                df_all = pd.concat(
                    [
                        df_all,
                        pd.DataFrame(
                            [
                                {
                                    "Дата": "",
                                    "Автор": "",
                                    "Тип оплаты": "",
                                    "Название": "ИТОГО",
                                    "Базовая цена": "",
                                    "Добавки": "",
                                    "Сумма добавок": "",
                                    "Сумма позиции": total_all,
                                    "Запрос": "",
                                    "Сотрудник": "",
                                }
                            ]
                        ),
                    ],
                    ignore_index=True,
                )
            df_all.to_excel(writer, sheet_name="Все позиции", index=False)

            if not df_prepared.empty and "Тип оплаты" in df_prepared:
                for pt in sorted(
                    df_prepared["Тип оплаты"].dropna().unique(), key=lambda s: str(s).lower()
                ):
                    mask = (
                        df_prepared["Тип оплаты"].astype(str).str.lower()
                        == str(pt).lower()
                    )
                    df_pt = df_prepared[mask].copy()
                    sheet = str(pt).capitalize()

                    total_pt = (
                        df_pt["Сумма позиции"].sum() if "Сумма позиции" in df_pt else 0
                    )
                    if not df_pt.empty:
                        df_pt = pd.concat(
                            [
                                df_pt,
                                pd.DataFrame(
                                    [
                                    {
                                        "Дата": "",
                                        "Автор": "",
                                        "Тип оплаты": sheet,
                                        "Название": "ИТОГО",
                                        "Базовая цена": "",
                                        "Добавки": "",
                                        "Сумма добавок": "",
                                        "Сумма позиции": total_pt,
                                        "Запрос": "",
                                        "Сотрудник": "",
                                    }
                                ]
                            ),
                        ],
                        ignore_index=True,
                        )
                    df_pt.to_excel(writer, sheet_name=sheet, index=False)

            if not df_prepared.empty:
                grouped = (
                    df_prepared.groupby(["Тип оплаты", "Название"], dropna=False)
                    .agg(
                        Количество=("Сумма позиции", "size"),
                        Общая_сумма=("Сумма позиции", "sum"),
                    )
                    .reset_index()
                )
            else:
                grouped = pd.DataFrame(
                    columns=["Тип оплаты", "Название", "Количество", "Общая_сумма"]
                )

            if not grouped.empty:
                grouped_total = pd.DataFrame(
                    [
                        [
                            "",
                            "ИТОГО",
                            grouped["Количество"].sum(),
                            grouped["Общая_сумма"].sum(),
                        ]
                    ],
                    columns=["Тип оплаты", "Название", "Количество", "Общая_сумма"],
                )
                grouped = pd.concat([grouped, grouped_total], ignore_index=True)

            grouped.to_excel(writer, sheet_name="Группировка", index=False)

            if not df_prepared.empty and "Автор" in df_prepared:
                by_author = (
                    df_prepared.groupby(["Автор"], dropna=False)
                    .agg(
                        Количество=("Сумма позиции", "size"),
                        Общая_сумма=("Сумма позиции", "sum"),
                    )
                    .reset_index()
                    .sort_values(["Общая_сумма"], ascending=False)
                )
                by_author.to_excel(writer, sheet_name="По авторам", index=False)

        auto_adjust_columns(path)

    orders_regular = orders_df[orders_df["is_staff"] == 0].copy()
    orders_staff = orders_df[orders_df["is_staff"] == 1].copy()

    prepared_regular = _prepare_orders_df(orders_regular)
    prepared_staff = _prepare_orders_df(orders_staff)

    if start_date and end_date:
        period_str = (
            start_date.isoformat()
            if start_date == end_date
            else f"{start_date.isoformat()}__{end_date.isoformat()}"
        )
    else:
        period_str = "all"

    report_path = f"report_{period_str}.xlsx"
    staff_report_path = (
        f"report_staff_{period_str}.xlsx" if not orders_staff.empty else None
    )
    log_path = f"log_report_{period_str}.xlsx"

    _write_report(prepared_regular, report_path)
    if staff_report_path:
        _write_report(prepared_staff, staff_report_path)

    # 7) Лог действий
    actions_df.columns = [
        "Дата/время",
        "Действие",
        "Тип оплаты",
        "Название",
        "user_id",
        "username",
        "Сотрудник",
    ]
    actions_df["Сотрудник"] = actions_df["Сотрудник"].apply(lambda v: bool(v))

    with pd.ExcelWriter(log_path, engine="openpyxl") as writer:
        actions_df.to_excel(writer, sheet_name="Журнал действий", index=False)

    auto_adjust_columns(log_path)

    return report_path, staff_report_path, log_path
